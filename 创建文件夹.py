import os
from openpyxl import Workbook, load_workbook

def mkdir(path):
    folder = os.path.exists(path)
    if not folder:
        os.makedirs(path)

#读取文件
path = os.getcwd()
print(path)
fileList = os.listdir(path)
fileName: str = None

n = 0
for i in fileList:
    if ".xlsx" in fileList[n]:
        fileName = fileList[n]
        break
    n += 1

wb = load_workbook(fileName, read_only=True)
my_sheet = wb.active
cell = tuple(my_sheet)

for i in range(1, my_sheet.max_row):
    pathend = os.getcwd() + "/" + str(cell[i][0].value) + "-" + str(cell[i][4].value) + "/"
    print(pathend)
    mkdir(pathend);
