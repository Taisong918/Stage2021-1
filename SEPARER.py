import os
from openpyxl import Workbook, load_workbook

path = os.getcwd()
fileList = os.listdir(path)
fileName: str = None

n = 0
for i in fileList:
    if "bisLISTE  STE VAT 最终版" in fileList[n]:
        fileName = fileList[n]
        break
    n += 1

wb = load_workbook(fileName, read_only=True)
my_sheet = wb.active
cell = tuple(my_sheet)

# print(my_sheet.max_row)
# print(my_sheet.max_column)

sie_Paris2 = Workbook()
sie_Paris2s = sie_Paris2.active
sie_Paris16 = Workbook()
sie_Paris16s = sie_Paris16.active
sie_BonneVille = Workbook()
sie_BonneVilles = sie_BonneVille.active
sie_Annency = Workbook()
sie_Annencys = sie_Annency.active
sie_Vide = Workbook()
sie_Vides = sie_Vide.active

row = ["Numéro", "Agence", "", "", "", "", "SIRENE", "", "", "", "", "", "", "", "", "SIRENE", "", "", "", "", "NOM CN"]
sie_Paris2s.append(row)
sie_Paris16s.append(row)
sie_BonneVilles.append(row)
sie_Annencys.append(row)
sie_Vides.append(row)

j = 13  # SIE
for i in range(my_sheet.max_row):
    if cell[i][j].value == "SIE PARIS 2EME":
        row[0] = cell[i][0].value
        row[1] = cell[i][1].value
        row[6] = cell[i][15].value[:9]
        row[15] = cell[i][15].value[:9]
        row[20] = cell[i][20].value
        sie_Paris2s.append(row)

    if cell[i][j].value == "SIE PARIS 16 CHAILLOT":
        row[0] = cell[i][0].value
        row[1] = cell[i][1].value
        row[6] = cell[i][15].value[:9]
        row[15] = cell[i][15].value[:9]
        row[20] = cell[i][20].value
        sie_Paris16s.append(row)

    if cell[i][j].value == "SIE BONNEVILLE":
        row[0] = cell[i][0].value
        row[1] = cell[i][1].value
        row[6] = cell[i][15].value[:9]
        row[15] = cell[i][15].value[:9]
        row[20] = cell[i][20].value
        sie_BonneVilles.append(row)

    if cell[i][j].value == "SIE ANNECY":
        row[0] = cell[i][0].value
        row[1] = cell[i][1].value
        row[6] = cell[i][15].value[:9]
        row[15] = cell[i][15].value[:9]
        row[20] = cell[i][20].value
        sie_Annencys.append(row)

    if cell[i][j].value == None and cell[i][15].value != None:
        row[0] = cell[i][0].value
        row[1] = cell[i][1].value
        row[6] = cell[i][15].value[:9]
        row[15] = cell[i][15].value[:9]
        row[20] = cell[i][20].value
        sie_Vides.append(row)
wb.close()

sie_Paris2.save("SIE PARIS 2EME.xlsx")
sie_Paris16.save("SIE PARIS 16 CHAILLOT.xlsx")
sie_BonneVille.save("SIE BONNEVILLE.xlsx")
sie_Annency.save("SIE ANNECY.xlsx")
sie_Vide.save("SIE VIDE.xlsx")