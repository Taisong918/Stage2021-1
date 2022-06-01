import time

import tkinter as tk
from openpyxl import Workbook, load_workbook
import win32api


def m0ee0():
    M0 = ["DENOMINATION 1",  # 0
          "DENOMINATION 2",  # 1
          "ADRESSE DU SIEGE Rés bât app étage n voie lieudit 1",  # 2
          "ADRESSE DU SIEGE Rés bât app étage n voie lieudit 2",  # 3
          "Commune",  # 4
          "DATE DE DÉBUT DACTIVITÉ",  # 5
          "Le",  # 6
          "Document Nom M0"  # 7
          ]
    # ------------------------------------------------------------------#
    EE0 = ["2DENOMINATION 1",  # 0
           "2DENOMINATION 2",  # 1
           "2Adresse rés bât n voie lieudit",  # 2
           "2Adresse rés bât n voie lieudit_2",  # 3
           "Commune_3",  # 4
           "Pays_3",  # 5
           "DATE DE DEBUT DACTIVITE EN FRANCE",  # 6
           "Le",
           "Document Nom EE0 Paris 16",  # 8
           "Document Nom EE0 Bonneville"  # 9
           ]

    outputpage.append(M0 + EE0)

    for i in range(1, inputpage.max_row):
        if cell[i][3].value is None:
            break
        # DENOMINATION
        denomination = cell[i][3].value.upper()
        couper: int = statisique(denomination, dictm0d, 0.983)
        if len(cell[i][3].value) >= couper:
            M0[0] = denomination[:couper]
            M0[1] = denomination[couper:len(cell[i][3].value)]
        else:
            M0[0] = denomination
            M0[1] = ''
        # ADRESSE DU SIEGE
        if len(cell[i][4].value) >= 24:
            M0[2] = cell[i][4].value[:23]
            M0[3] = cell[i][4].value[23:len(cell[i][4].value)]
        else:
            M0[2] = cell[i][4].value
            M0[3] = ''
        # COMMUNE
        M0[4] = str(cell[i][5].value) + " " + cell[i][6].value + " " + cell[i][7].value + " " + cell[i][8].value
        # DATE DE DÉBUT DACTIVITÉ
        M0[5] = cell[i][10].value.replace('/', '')
        # LE
        if cell[i][11].value == "当天":
            M0[6] = time.strftime("%d/%m/%Y", time.localtime())
        else:
            M0[6] = cell[i][11].value
        # DOCUMENT NOM
        M0[7] = str(cell[i][0].value) + "-M0-" + cell[i][3].value.upper()
        # ------------------------------------------------------------------#
        # DENOMINATION
        denomination = cell[i][3].value.upper()
        couper: int = statisique(denomination, dictm0d, 0.71)
        if len(cell[i][3].value) >= couper:
            EE0[0] = denomination[:couper]
            EE0[1] = denomination[couper:len(cell[i][3].value)]
        else:
            EE0[0] = denomination
            EE0[1] = ''
        # ADRESSE
        if len(cell[i][4].value) >= 37:
            EE0[2] = cell[i][4].value.lower()[:36]
            EE0[3] = cell[i][4].value.lower()[36:len(cell[i][4].value)]
        else:
            EE0[2] = cell[i][4].value.lower()
            EE0[3] = ''
        # COMMUNE
        EE0[4] = str(cell[i][5].value) + " " + cell[i][6].value + " " + cell[i][7].value
        # Pays
        EE0[5] = cell[i][8].value
        # DATE DE DEBUT DACTIVITE EN FRANCE
        EE0[6] = M0[5]
        # LE
        EE0[7] = M0[6]
        # Document Nom
        # 16
        EE0[8] = cell[i][3].value.upper() + "-EE0-" + str(cell[i][0].value)
        # Bonneville
        EE0[9] = str(cell[i][0].value) + "-EE0-" + cell[i][3].value.upper()

        outputpage.append(M0 + EE0)

    output.save("预处理\\预处理.xlsx")
    input.close()
    output.close()
    win32api.ShellExecute(0, 'open', 'C:\\AutoFill\\BulkPDF\\BulkPDF.exe', '', '', 1)
    window.quit()


def m2():
    M2 = ["N UNIQUE DIDENTIFICATION",  # 0
          "Dénomination  Sigle",  # 1
          "Forme juridique",  # 2
          "LETABLISSEMENT DEVIENT",  # 3
          "Le",  # 4
          "Nom",  # 5
          ]

    outputpage.append(M2)

    for i in range(1, inputpage.max_row):
        if cell[i][10].value is None:
            break

        # N UNIQUE DIDENTIFICATION
        M2[0] = str(cell[i][10].value)

        denomination = str(cell[i][7].value).upper()
        couper: int = statisique(denomination, dictm0d, 1)
        if len(cell[i][7].value) >= couper:
            M2[1] = denomination[:couper]
            M2[2] = denomination[couper:len(cell[i][7].value)]
        else:
            M2[1] = denomination
            M2[2] = ''

        M2[3] = str(cell[i][2].value).replace('/', '')
        M2[4] = str(cell[i][3].value).replace('/', '')
        M2[5] = str(cell[i][0].value) + '-M2-' + str(cell[i][7].value).upper()
        outputpage.append(M2)

    output.save("预处理\\预处理.xlsx")
    input.close()
    output.close()
    win32api.ShellExecute(0, 'open', 'C:\\AutoFill\\BulkPDF\\BulkPDF.exe', '', '', 1)
    window.quit()


def m4():
    M4: dict = [
        "1",  # 0
        "7",  # 1
        "8",  # 2
        "11",  # 3
        "62",  # 4
        "Nom"  # 5
    ]

    outputpage.append(M4)

    for i in range(1, inputpage.max_row):
        if cell[i][0].value is None:
            break
        M4[0] = str(cell[i][6].value)
        M4[1] = str(cell[i][3].value)
        M4[2] = "Error"
        if "02" in str(cell[i][9].value):
            M4[2] = "LOGEFI SERVICES  12 RUE VIVIENNE  75002  PARIS"
        if "16" in str(cell[i][9].value):
            M4[2] = "MORGAN TAX  112 AV KLEBER 75116 PARIS"
        if "ille" in str(cell[i][9].value):
            M4[2] = "JACOB ADVISORY 160 A route de l’Arroz, 74300 Châtillon-sur-Cluses"
        M4[3] = str(cell[i][7].value).replace('/', '')
        M4[4] = str(cell[i][8].value)
        M4[5] = str(cell[i][0].value) + "-M4-" + str(cell[i][6].value)
        outputpage.append(M4)

    output.save("预处理\\预处理.xlsx")
    input.close()
    output.close()
    win32api.ShellExecute(0, 'open', 'C:\\AutoFill\\BulkPDF\\BulkPDF.exe', '', '', 1)
    window.quit()


def statisique(c: str, d: dict, para: float):
    somme: float = 0
    resultat: int = 99999
    for i in range(0, len(c)):
        somme += d.get(c[i], float(1 / 60))
        if somme >= para:
            resultat = i + 1
            break
    return resultat


dictm0d = {
    'A': float(1 / 36),
    'B': float(1 / 36),
    'C': float(1 / 33),
    'D': float(1 / 33),
    'E': float(1 / 36),
    'F': float(1 / 39),
    'G': float(1 / 31),
    'H': float(1 / 33),
    'I': float(1 / 84),
    'J': float(1 / 48),
    'K': float(1 / 36),
    'L': float(1 / 43),
    'M': float(1 / 29),
    'N': float(1 / 33),
    'O': float(1 / 31),
    'P': float(1 / 36),
    'Q': float(1 / 31),
    'R': float(1 / 33),
    'S': float(1 / 36),
    'T': float(1 / 39),	
    'U': float(1 / 33),
    'V': float(1 / 36),
    'W': float(1 / 25),
    'X': float(1 / 36),
    'Y': float(1 / 36),
    'Z': float(1 / 39),
    ' ': float(1 / 86),
}

input = load_workbook("模板\\模板.xlsx")
inputpage = input.active

output = Workbook()
outputpage = output.active

cell = tuple(inputpage)

window = tk.Tk()
window.title('AutoFill')
window.geometry('130x100')
m0ee0 = tk.Button(window, text='输出M0或者EE0', command=m0ee0)
m0ee0.pack()
m2 = tk.Button(window, text='输出M2', command=m2)
m2.pack()
m4 = tk.Button(window, text='输出M4', command=m4)
m4.pack()

window.mainloop()
