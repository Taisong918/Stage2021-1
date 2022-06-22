import os
import shutil
import time
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

class CreateFolder:
    file_list: list[str]
    path_pdf: str
    dict_agence: dict
    date: str
    path: str
    wb: Workbook
    my_sheet: Worksheet
    cell: tuple

    def __init__(self, path):
        #初始化日期
        self.date = time.strftime("%Y.%m.%d", time.localtime())
        #pdf文件
        self.path_pdf = path
        #读取文件
        self.path = path
        self.file_list = os.listdir(self.path)
        file_name: str = ""
        for file in self.file_list:
            if ".xlsx" in file:
                file_name = file
                break
        self.wb = load_workbook(file_name, read_only=True)
        self.my_sheet = self.wb.active
        self.cell = tuple(self.my_sheet)
        #统计AGENCE数量
        self.dict_agence = {}
        for i in range(1, self.my_sheet.max_row):
            key = str(self.cell[i][1].value)
            value = self.dict_agence.get(key)
            if value is None:
                self.dict_agence[key] = 1
            else:
                self.dict_agence[key] = value + 1
            #出现空白表格跳出
            if self.cell[i][1].value is None:
                break

    def mkdir(self, path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)

    def remove_suffix(self):
        for file in self.file_list:
            if "_CompressPdf" in file:
                os.rename(self.path_pdf + os.sep + file, self.path_pdf + os.sep + file.replace("_CompressPdf", ""))
        self.file_list = os.listdir(self.path)

    def create_folder(self):
        #1 第一层
        folder_name = self.date + " Paris 02 " + str(self.my_sheet.max_row - 1) + "家" #去除第一行
        self.mkdir(folder_name)
        self.path = self.path + os.sep + folder_name
        #2 第二层
        for k in self.dict_agence.keys():
            v = self.dict_agence.get(k)
            path_inner = self.path + os.sep + str(self.date) + " " + str(k) + " " + str(v) + "家"
            self.mkdir(path_inner)
            #3 第三层
            self.create_folder_sup(k, v, path_inner)

    def create_folder_sup(self, agence_name, agence_nbr, path_inner):
        cpt = 1
        while 0 < agence_nbr and cpt < self.my_sheet.max_row:
            if self.cell[cpt][1].value == agence_name:
                folder_name = str(self.cell[cpt][0].value) + " " + self.cell[cpt][3].value.upper()
                path_inner_inner = path_inner + os.sep + folder_name
                self.mkdir(path_inner_inner)
                self.move_docs(path_inner_inner, str(self.cell[cpt][0].value)) #prefix
                agence_nbr -= 1
            cpt += 1

    def move_docs(self, path_inner_inner: str, prefix: str):
        for file in self.file_list:
            if ".pdf" in file and prefix in file:
                shutil.move(self.path_pdf + os.sep + file, path_inner_inner + os.sep + file)


    def create_xlsx_doc(self):
        doc = Workbook()
        docs = doc.active
        row = ["N°", "Date d’envoi", "Nom", "Code Postal", "Ville", "Pays","Date de début d’acitivité",
               "Activités", "Nom du correspondant", "Complément d’adresse", "Adresse", "Date de signature", "TVAs"]
        docs.append(row)
        row[1] = time.strftime("%d/%m/%Y", time.localtime())
        row[5] = "Chine"
        row[7] = "Vente en ligne"
        row[8] = "Li Zheng"
        row[9] = "Chez LOGEFI Services"
        row[10] = "12 rue Vivienne"
        row[12] = "Trimestrielle"

        cpt = 1
        for i in range(1, self.my_sheet.max_row):
            row[0] = cpt
            row[2] = self.cell[i][3].value
            row[3] = str(self.cell[i][5].value)
            row[4] = self.cell[i][6].value + " " + self.cell[i][7].value
            row[6] = str(self.cell[i][10].value)
            row[11] = str(self.cell[i][11].value)
            docs.append(row)
            cpt += 1
            # 出现空白表格跳出
            if self.cell[i][1].value is None:
                break
        date = time.strftime("%d.%m.%Y", time.localtime())
        doc.save(self.path + os.sep + date + " Créations M0 - " + str(self.my_sheet.max_row - 1) + "sociétés.xlsx")

file_name = os.getcwd()
cf = CreateFolder(file_name)
cf.remove_suffix()
cf.create_folder()
cf.create_xlsx_doc()