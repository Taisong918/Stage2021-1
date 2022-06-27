import os
from openpyxl import Workbook, load_workbook

#读取文件
path = os.getcwd()
fileList = os.listdir(path)
fileName: str = None

n = 0
for i in fileList:
    if ".xlsx" in fileList[n]:
        fileName = fileList[n]
        break
    n += 1

wb = load_workbook(fileName, read_only=True)
my_sheet = wb.active
cell = tuple(my_sheet)

#获取代理商名称
s = set()
for i in range(1,my_sheet.max_row):
    s.add(str(cell[i][1].value))

#查找保存
def stockageDesDocuments(name: str):
    w = Workbook()
    ws = w.active
    ws.append(["numéro","Agence","公司名","SIREN","未申报月份","Précisions"])
    for i in range(my_sheet.max_row):
        if str(cell[i][1].value) == name:
            row_data = []
            for j in range(0, my_sheet.max_column):
                row_data.append(cell[i][j].value)
            ws.append(row_data)
    w.save(name + ".xlsx")

for name in iter(s):
    stockageDesDocuments(str(name))
