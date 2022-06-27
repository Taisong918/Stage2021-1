import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class Separate:
    path: str
    wb: Workbook
    my_sheet: Worksheet
    cell: tuple
    row: list

    def __init__(self):
        self.path = os.getcwd()
        file_list = os.listdir(self.path)
        file_name: str = ""
        for file in file_list:
            if "bisLISTE  STE VAT 最终版" in file:
                file_name = file
                break
        self.wb = load_workbook(file_name, read_only=True)
        self.my_sheet = self.wb.active
        self.cell = tuple(self.my_sheet)

    def __init__(self, output_path, file_name):
        self.path = output_path
        self.wb = load_workbook(file_name, read_only=True)
        self.my_sheet = self.wb.active
        self.cell = tuple(self.my_sheet)

    def classification(self):
        row = ["Numéro", "Agence", "", "", "", "", "SIRENE", "", "", "", "", "", "", "", "", "SIRENE", "", "", "", "",
               "NOM CN"]

        sie_paris2 = Workbook()
        sie_paris2s = sie_paris2.active
        sie_paris16 = Workbook()
        sie_paris16s = sie_paris16.active
        sie_bonneville = Workbook()
        sie_bonnevilles = sie_bonneville.active
        sie_annency = Workbook()
        sie_annencys = sie_annency.active
        sie_vide = Workbook()
        sie_vides = sie_vide.active
        sie_autre = Workbook()
        sie_autres = sie_autre.active

        sie_paris2s.append(row)
        sie_paris16s.append(row)
        sie_bonnevilles.append(row)
        sie_annencys.append(row)
        sie_vides.append(row)
        sie_autres.append(row)

        j = 13
        for i in range(1, self.my_sheet.max_row):
            if self.cell[i][15].value is not None:
                row[0] = self.cell[i][0].value
                row[1] = self.cell[i][1].value
                row[6] = self.cell[i][15].value[:9]
                row[15] = self.cell[i][15].value[:9]
                row[20] = self.cell[i][20].value

                match self.cell[i][j].value:
                    case "SIE PARIS 2EME":
                        sie_paris2s.append(row)
                    case "SIE PARIS 16 CHAILLOT":
                        sie_paris16s.append(row)
                    case "SIE ANNECY":
                        sie_annencys.append(row)
                    case "SIE BONNEVILLE":
                        sie_bonnevilles.append(row)
                    case None:
                        sie_vides.append(row)
                    case _:
                        sie_autres.append(row)

        self.storing_documents(self.path + os.sep + "SIE PARIS 2EME.xlsx", sie_paris2)
        self.storing_documents(self.path + os.sep + "SIE PARIS 16 CHAILLOT.xlsx", sie_paris16)
        self.storing_documents(self.path + os.sep + "SIE BONNEVILLE.xlsx", sie_bonneville)
        self.storing_documents(self.path + os.sep + "SIE ANNECY.xlsx", sie_annency)
        self.storing_documents(self.path + os.sep + "SIE VIDE.xlsx", sie_vide)
        self.storing_documents(self.path + os.sep + "SIE AUTRE.xlsx", sie_autre)

    def storing_documents(self, name: str, w: Workbook):
        w.save(name)


#Separate().classification()
